from openpyxl import load_workbook

wb = load_workbook(filename='template base de datos.xlsx')

wb.get_sheet_names()
sheet1 = wb.get_sheet_by_name('Venezuela')
sheet2 = wb.get_sheet_by_name('Peru')
sheet3 = wb.get_sheet_by_name('Viajes')
sheet4 = wb.get_sheet_by_name('Reporte final')

print("¡Bienvenido Carlos!. ¿En qué te puedo ayudar?")

print()

opciones_de_entrada = (
"1 -> Consultar el gasto asociado a uno o varios servicios correspondientes a uno o varios meses",
"2 -> Consultar el gasto total de todos los servicios en un mes particular",
"3 -> Consultar el gasto total anual de un servicio particular",
"4 -> Consultar reporte del total a pagar")

for i in range(len(opciones_de_entrada)):
    print(opciones_de_entrada[i])

print()

opcion_seleccionada = int(input("Introduce el número de tu opción seleccionada: "))

if opcion_seleccionada == 1:
    pais = input("Indique pais, Venezuela o Peru: ")
    if pais == sheet1.title:
        n_servicios_vzla = int(input("Ingrese cantidad de servicios a consultar: "))
        lista1 = []
        for i in range(n_servicios_vzla):
            servicios_vzla = input("Ingrese servicio #" + str(i+1) + ": ")
            lista1.append(servicios_vzla)
        n_meses_vzla = int(input("Ingrese cantidad de meses a consultar: "))
        lista2 = []
        for i in range(n_meses_vzla):
            meses_vzla = input("Ingrese mes #" + str(i+1) + ": ")
            lista2.append(meses_vzla)
        lista3 = []
        for item in lista1:
            for i in range(3, 18, 1):
                if item == sheet1.cell(row=1, column = i).internal_value:
                    for mes in lista2:
                        for j in range(2, 14, 1):
                            if mes == sheet1.cell(row=j, column=2).internal_value:
                                lista3.append(sheet1.cell(row=j, column = i).internal_value)
        h = 0
        for i in range(n_servicios_vzla):
            for j in range(n_meses_vzla):
                print("Para el servicio #" + str(i+1) + " en el mes #" + str(j+1) + ", el gasto es: ", lista3[h])
                h += 1
    if pais == sheet2.title:
        n_servicios_peru = int(input("Ingrese cantidad de servicios a consultar: "))
        lista1 = []
        for i in range(n_servicios_peru):
            servicios_peru = input("Ingrese servicio #" + str(i+1) + ": ")
            lista1.append(servicios_peru)
        n_meses_peru = int(input("Ingrese cantidad de meses a consultar: "))
        lista2 = []
        for i in range(n_meses_peru):
            meses_peru = input("Ingrese mes #" + str(i+1) + ": ")
            lista2.append(meses_peru)
        lista3 = []
        for item in lista1:
            for i in range(3, 13, 1):
                if item == sheet2.cell(row=1, column = i).internal_value:
                    for mes in lista2:
                        for j in range(2, 14, 1):
                            if mes == sheet2.cell(row=j, column=2).internal_value:
                                lista3.append(sheet2.cell(row=j, column = i).internal_value)
        h = 0
        for i in range(n_servicios_peru):
            for j in range(n_meses_peru):
                print("Para el servicio #" + str(i+1) + " en el mes #" + str(j+1) + ", el gasto es: ", lista3[h])
                h += 1

if opcion_seleccionada == 2:
    pais = input("Indique pais, Venezuela o Peru: ")
    if pais == sheet1.title:
        for i in range(2, 14, 1):
            print(sheet1.cell(row=i, column=2).internal_value)
        mes_seleccionado_vzla = input("Indique el mes: ")
        for i in range(2, 14, 1):
            if mes_seleccionado_vzla == sheet1.cell(row=i, column=2).internal_value:
                print(sheet1.cell(row=i, column=18).value)
    if pais == sheet2.title:
        for i in range(2, 14, 1):
            print(sheet2.cell(row=i, column=2).internal_value)
        mes_seleccionado_peru = input("Indique el mes: ")
        for i in range(2, 14, 1):
            if mes_seleccionado_peru == sheet2.cell(row=i, column=2).internal_value:
                print(sheet2.cell(row=i, column=13).value)

if opcion_seleccionada == 3:
    pais = input("Indique pais, Venezuela o Peru: ")
    if pais == sheet1.title:
        for i in range(3, 18, 1):
            print(sheet1.cell(row=1, column=i).internal_value)
        servicio_seleccionado_vzla = input("Indique el servicio: ")
        for i in range(3, 18, 1):
            if servicio_seleccionado_vzla == sheet1.cell(row=1, column=i).internal_value:
                print(sheet1.cell(row=14, column=i).value)
    if pais == sheet2.title:
        for i in range(3, 13, 1):
            print(sheet2.cell(row=1, column=i).internal_value)
        servicio_seleccionado_peru = input("Indique el servicio: ")
        for i in range(3, 13, 1):
            if servicio_seleccionado_peru == sheet2.cell(row=1, column=i).internal_value:
                print(sheet2.cell(row=14, column=i).value)

if opcion_seleccionada == 4:
    for row in sheet4.iter_rows():
        for i, cell in enumerate(row):
            value = " "
            if cell.internal_value is not None:
                value = cell.internal_value
            if type(str()) == type(value):
                print("{:^15s} ".format(value), end="")
            elif type(value) in [type(int()), type(float())] and i != 0:
                print("{:15,.2f} ".format(value), end="")
            elif type(value) in [type(int()), type(float())] and i == 0:
                print("{:^15s} ".format(str(value)), end="")
            else:
                print(value, type(value))
        print(" ")
